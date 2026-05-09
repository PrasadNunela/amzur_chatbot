"""Attachment processing service — extract content from uploaded files.

General-purpose approach:
- Try to detect file type from extension, MIME type, and file content
- Apply appropriate extraction method, with graceful fallbacks
- Never assume file structure - preserve data as-is
- Handle unknown formats without breaking
"""

import base64
import mimetypes
from pathlib import Path
from uuid import UUID
import csv
import json
import cv2
from openpyxl import load_workbook

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.chat import Attachment


class AttachmentService:
    """Service for processing and extracting content from attachments.
    
    Uses a format-agnostic approach:
    1. Detect file type from extension, MIME type, and content
    2. Try specialized extractors for known formats
    3. Fall back to text extraction for unknown formats
    4. Always return usable content without breaking
    """

    @staticmethod
    async def get_attachment(db: AsyncSession, attachment_id: UUID) -> Attachment | None:
        """Retrieve an attachment by ID."""
        stmt = select(Attachment).where(Attachment.id == attachment_id)
        result = await db.execute(stmt)
        return result.scalar()

    @staticmethod
    def _get_full_path(file_path: str) -> Path:
        """Get full absolute path for a file, handling both relative and absolute paths."""
        from app.core.config import settings
        
        path = Path(file_path)
        
        # If already absolute, use it as-is
        if path.is_absolute():
            return path
        
        # If relative, construct from UPLOAD_DIR
        upload_dir = Path(settings.UPLOAD_DIR).resolve()  # Resolve to absolute path
        full_path = upload_dir / path
        return full_path.resolve()  # Normalize the path

    @staticmethod
    def extract_text_content(file_path: str) -> str:
        """Extract text content from a file.
        
        Supports:
        - Plain text files (txt, json, xml, code)
        - PDF (basic text extraction)
        - CSV (raw content)
        """
        from app.core.logger import logger
        
        path = AttachmentService._get_full_path(file_path)
        
        if not path.exists():
            logger.warning(f"Text file not found: {path}")
            return ""
        
        try:
            # For text-based files
            content = path.read_text(encoding='utf-8', errors='ignore')
            logger.info(f"Successfully extracted {len(content)} characters from {path.name}")
            return content
        except Exception as e:
            logger.error(f"Failed to extract text from {path}: {e}")
            return ""

    @staticmethod
    def extract_xlsx_content(file_path: str) -> str:
        """Extract content from Excel XLSX file as markdown table.
        
        Preserves all values as-is without format assumptions.
        Let the AI interpret the data semantics.
        
        Args:
            file_path: Path to the XLSX file
            
        Returns:
            Markdown formatted table representation of the spreadsheet
        """
        from app.core.logger import logger
        
        path = AttachmentService._get_full_path(file_path)
        
        if not path.exists():
            logger.warning(f"XLSX file not found: {path}")
            return ""
        
        try:
            workbook = load_workbook(str(path))
            all_content = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Get all rows with at least one non-empty cell
                rows = []
                for row in sheet.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        rows.append(row)
                
                if not rows:
                    continue
                
                # Add sheet name as header if there are multiple sheets
                if len(workbook.sheetnames) > 1:
                    all_content.append(f"## Sheet: {sheet_name}\n")
                
                # Convert all cells to strings (preserve values as-is)
                formatted_rows = []
                for row in rows:
                    formatted_row = []
                    for cell in row:
                        if cell is None:
                            formatted_row.append("")
                        else:
                            # Convert to string, preserve the exact value
                            cell_str = str(cell).strip()
                            formatted_row.append(cell_str)
                    formatted_rows.append(formatted_row)
                
                # Use first row as header
                if formatted_rows:
                    header = formatted_rows[0]
                    data_rows = formatted_rows[1:] if len(formatted_rows) > 1 else []
                    
                    # Build markdown table
                    all_content.append("| " + " | ".join(header) + " |\n")
                    all_content.append("|" + "|".join([" --- "] * len(header)) + "|\n")
                    
                    # Add data rows
                    for row in data_rows:
                        # Pad row to match header length
                        while len(row) < len(header):
                            row.append("")
                        all_content.append("| " + " | ".join(row[:len(header)]) + " |\n")
                
                all_content.append("\n")
            
            workbook.close()
            result = "".join(all_content).strip()
            logger.info(f"Successfully extracted XLSX: {path.name} with {len(workbook.sheetnames)} sheets")
            return result
        except Exception as e:
            logger.error(f"Failed to extract XLSX from {path}: {e}")
            return ""


    @staticmethod
    def encode_image_base64(file_path: str) -> str:
        """Encode an image file as base64 for LLM input."""
        from app.core.logger import logger
        
        path = AttachmentService._get_full_path(file_path)
        
        if not path.exists():
            logger.warning(f"Image file not found: {path}")
            return ""
        
        try:
            file_content = path.read_bytes()
            b64 = base64.b64encode(file_content).decode('utf-8')
            logger.info(f"Successfully encoded image {path.name} ({len(file_content)} bytes)")
            return b64
        except Exception as e:
            logger.error(f"Failed to encode image {path}: {e}")
            return ""

    @staticmethod
    def extract_csv_content(file_path: str) -> str:
        """Extract content from CSV file as markdown table.
        
        Preserves all values as-is.
        
        Args:
            file_path: Path to the CSV file
            
        Returns:
            Markdown formatted table
        """
        from app.core.logger import logger
        import csv
        
        path = AttachmentService._get_full_path(file_path)
        
        if not path.exists():
            logger.warning(f"CSV file not found: {path}")
            return ""
        
        try:
            with open(path, 'r', encoding='utf-8', errors='ignore') as f:
                reader = csv.reader(f)
                rows = list(reader)
            
            if not rows:
                logger.warning(f"CSV file is empty: {path}")
                return ""
            
            # Build markdown table
            content = []
            for row_idx, row in enumerate(rows):
                # Skip completely empty rows
                if not any(cell.strip() for cell in row):
                    continue
                
                row_content = "| " + " | ".join(cell.strip() for cell in row) + " |\n"
                content.append(row_content)
                
                # Add separator after first row (header)
                if row_idx == 0:
                    separator = "|" + "|".join([" --- "] * len(row)) + "|\n"
                    content.append(separator)
            
            result = "".join(content).strip()
            logger.info(f"Successfully extracted CSV: {path.name}")
            return result
        except Exception as e:
            logger.error(f"Failed to extract CSV from {path}: {e}")
            return ""
    
    @staticmethod
    def extract_json_content(file_path: str) -> str:
        """Extract content from JSON file, formatted nicely.
        
        Args:
            file_path: Path to the JSON file
            
        Returns:
            Formatted JSON content as code block
        """
        from app.core.logger import logger
        import json
        
        path = AttachmentService._get_full_path(file_path)
        
        if not path.exists():
            logger.warning(f"JSON file not found: {path}")
            return ""
        
        try:
            with open(path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # Format with proper indentation
            formatted = json.dumps(data, indent=2)
            logger.info(f"Successfully extracted JSON: {path.name}")
            return formatted
        except Exception as e:
            logger.error(f"Failed to extract JSON from {path}: {e}")
            return ""

    @staticmethod
    def extract_video_frames(file_path: str, num_frames: int = 5) -> list[str]:
        """Extract key frames from a video file and encode as base64 JPEG.
        
        Args:
            file_path: Path to the video file
            num_frames: Number of frames to extract (default 5)
        
        Returns:
            List of base64-encoded JPEG images
        """
        from app.core.logger import logger
        
        path = AttachmentService._get_full_path(file_path)
        
        if not path.exists():
            logger.warning(f"Video file not found: {path}")
            return []
        
        video = None
        try:
            video = cv2.VideoCapture(str(path))
            if not video.isOpened():
                logger.error(f"Failed to open video file: {path}")
                return []
            
            total_frames = int(video.get(cv2.CAP_PROP_FRAME_COUNT))
            if total_frames == 0:
                logger.warning(f"Video has no frames: {path}")
                return []
            
            logger.info(f"Extracting {num_frames} frames from video {path.name} ({total_frames} total frames)")
            
            # Calculate frame intervals to distribute extraction evenly
            frame_indices = [int(i * total_frames / num_frames) for i in range(num_frames)]
            
            frames_b64 = []
            for frame_idx in frame_indices:
                video.set(cv2.CAP_PROP_POS_FRAMES, frame_idx)
                ret, frame = video.read()
                
                if not ret or frame is None:
                    continue
                
                # Resize frame to reduce data size (max 1280 width)
                height, width = frame.shape[:2]
                if width > 1280:
                    ratio = 1280 / width
                    new_height = int(height * ratio)
                    frame = cv2.resize(frame, (1280, new_height))
                
                # Convert BGR to RGB and encode as JPEG
                frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                success, buffer = cv2.imencode('.jpg', frame_rgb, [cv2.IMWRITE_JPEG_QUALITY, 85])
                
                if success:
                    b64_string = base64.b64encode(buffer).decode('utf-8')
                    frames_b64.append(b64_string)
            
            logger.info(f"Successfully extracted {len(frames_b64)} frames from video")
            return frames_b64
            
        except Exception:
            return []
        finally:
            if video is not None:
                video.release()

    @staticmethod
    def get_attachment_content(attachment: Attachment) -> str | dict | list:
        """Extract and format attachment content for LLM input.
        
        Universal approach that handles ANY file format:
        1. Detect file type from extension and MIME type
        2. Try specialized extractors for known formats
        3. Fall back to generic text extraction
        4. As last resort, provide hex preview of unknown binary formats
        
        Always returns something useful - never breaks on unknown formats.
        
        Returns:
            - For images: dict with base64 content
            - For video: list of dicts with frame base64 content
            - For other types: str with formatted content
        """
        from app.core.logger import logger
        
        file_type = attachment.file_type
        mime_type = attachment.mime_type
        file_path = attachment.file_path
        filename = attachment.filename
        
        # Check if file exists
        full_path = AttachmentService._get_full_path(file_path)
        if not full_path.exists():
            logger.warning(f"Attachment file not found: {full_path} (original: {file_path})")
            return f"[Attachment: {filename} - file not found]"
        
        logger.info(f"Processing {file_type} attachment: {filename} from {full_path}")
        
        # Image handling (special case - return dict for vision models)
        if file_type == "image":
            base64_content = AttachmentService.encode_image_base64(file_path)
            if base64_content:
                return {
                    "type": "image",
                    "filename": filename,
                    "mime_type": mime_type,
                    "base64": base64_content,
                }
            return f"[Image: {filename} - could not read]"
        
        # Video handling (special case - return frames for vision models)
        if file_type == "video":
            frames = AttachmentService.extract_video_frames(file_path, num_frames=5)
            if frames:
                return {
                    "type": "video_frames",
                    "filename": filename,
                    "mime_type": mime_type,
                    "frames": frames,
                }
            return f"[Video: {filename} - could not extract frames]"
        
        # For all other file types, try extraction strategies in order
        content = None
        
        # Try format-specific extractors based on filename and MIME type
        filename_lower = filename.lower()
        
        # XLSX/Excel files
        if filename_lower.endswith(('.xlsx', '.xls')) or 'spreadsheet' in mime_type or 'excel' in mime_type:
            logger.debug(f"Attempting XLSX extraction for {filename}")
            content = AttachmentService.extract_xlsx_content(file_path)
            if content:
                return f"## {filename}\n\n{content}"
        
        # CSV files
        if filename_lower.endswith('.csv') or mime_type == 'text/csv':
            logger.debug(f"Attempting CSV extraction for {filename}")
            content = AttachmentService.extract_csv_content(file_path)
            if content:
                return f"## {filename}\n\n{content}"
        
        # JSON files
        if filename_lower.endswith('.json') or mime_type == 'application/json':
            logger.debug(f"Attempting JSON extraction for {filename}")
            content = AttachmentService.extract_json_content(file_path)
            if content:
                return f"```json\n{content}\n```"
        
        # Code files (based on extension)
        code_extensions = ('.py', '.js', '.ts', '.jsx', '.tsx', '.java', '.cpp', '.c', '.cs', '.sql', '.sh', '.yaml', '.yml')
        if filename_lower.endswith(code_extensions):
            logger.debug(f"Attempting code extraction for {filename}")
            content = AttachmentService.extract_text_content(file_path)
            if content:
                lang = AttachmentService._get_language_from_mime(mime_type)
                return f"```{lang}\n{content}\n```"
        
        # Generic text extraction (works for many text-based formats)
        logger.debug(f"Attempting generic text extraction for {filename}")
        content = AttachmentService.extract_text_content(file_path)
        if content:
            # Format based on file type hints
            if file_type == "code" or 'code' in filename_lower or 'script' in mime_type:
                lang = AttachmentService._get_language_from_mime(mime_type)
                return f"```{lang}\n{content}\n```"
            else:
                return f"## {filename}\n\n{content}"
        
        # Last resort: try generic extraction (handles binary files gracefully)
        logger.debug(f"Attempting generic fallback extraction for {filename}")
        fallback_content = AttachmentService._try_generic_extraction(file_path)
        if fallback_content:
            return f"## {filename}\n\n{fallback_content}"
        
        # Final fallback: provide file info
        try:
            file_size = full_path.stat().st_size
            return f"[File: {filename} ({file_size} bytes) - content could not be extracted]"
        except Exception:
            return f"[File: {filename} - could not access or read]"

    @staticmethod
    def _get_language_from_mime(mime_type: str) -> str:
        """Get code language identifier from MIME type."""
        mime_to_lang = {
            'text/x-python': 'python',
            'text/javascript': 'javascript',
            'text/typescript': 'typescript',
            'text/x-java': 'java',
            'text/x-csharp': 'csharp',
            'application/json': 'json',
            'application/xml': 'xml',
            'text/xml': 'xml',
            'text/plain': 'text',
            'text/csv': 'csv',
            'application/x-sql': 'sql',
            'text/x-shellscript': 'bash',
            'text/x-c': 'c',
            'text/x-c++': 'cpp',
        }
        return mime_to_lang.get(mime_type, 'text')
    
    @staticmethod
    def _try_generic_extraction(file_path: str) -> str | None:
        """Try to extract content from any file format as a last resort.
        
        Attempts multiple strategies:
        1. Read as UTF-8 text
        2. Read as Latin-1 (fallback for non-UTF-8)
        3. Hex dump with byte preview
        
        Returns:
            Extracted content or None if file cannot be read
        """
        from app.core.logger import logger
        
        full_path = AttachmentService._get_full_path(file_path)
        
        if not full_path.exists():
            return None
        
        # Get file size
        file_size = full_path.stat().st_size
        
        # Don't process extremely large files (>100MB)
        if file_size > 100 * 1024 * 1024:
            return f"[File too large to preview: {file_size / (1024*1024):.1f}MB]"
        
        try:
            # Try UTF-8
            with open(full_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read(10000)  # First 10KB
                if content.strip():
                    logger.info(f"Successfully extracted text from {full_path.name} (UTF-8)")
                    # If content is long, truncate with indicator
                    if len(content) == 10000:
                        content = content + "\n\n[... file continues ...]"
                    return content
        except Exception as e:
            logger.debug(f"UTF-8 extraction failed for {full_path.name}: {e}")
        
        try:
            # Try Latin-1 as fallback
            with open(full_path, 'r', encoding='latin-1', errors='ignore') as f:
                content = f.read(10000)  # First 10KB
                if content.strip():
                    logger.info(f"Successfully extracted text from {full_path.name} (Latin-1)")
                    if len(content) == 10000:
                        content = content + "\n\n[... file continues ...]"
                    return f"[Text content extracted (Latin-1 encoding)]:\n{content}"
        except Exception as e:
            logger.debug(f"Latin-1 extraction failed for {full_path.name}: {e}")
        
        try:
            # Try binary read with hex preview
            with open(full_path, 'rb') as f:
                binary_data = f.read(1000)  # First 1KB
                if binary_data:
                    hex_preview = binary_data[:200].hex()
                    ascii_preview = ''.join(
                        chr(b) if 32 <= b < 127 else '.'
                        for b in binary_data[:200]
                    )
                    return f"[Binary file - cannot extract text]\nSize: {file_size} bytes\nPreview (hex): {hex_preview}\nPreview (ASCII): {ascii_preview}"
        except Exception as e:
            logger.debug(f"Binary extraction failed for {full_path.name}: {e}")
        
        return None

    @staticmethod
    def _detect_file_type_with_fallback(filename: str, mime_type: str = '') -> str:
        """Detect file type intelligently using multiple strategies.
        
        Priority:
        1. Known extension mapping (fast and reliable)
        2. MIME type hints
        3. Default to 'document' for safe processing
        """
        ext = filename.lower().split('.')[-1] if '.' in filename else ''
        
        # Strategy 1: Extension mapping (most reliable)
        type_map = {
            # Spreadsheets
            'xlsx': 'table',
            'xls': 'table',
            'csv': 'table',
            'ods': 'table',
            
            # Documents
            'pdf': 'document',
            'docx': 'document',
            'doc': 'document',
            'txt': 'document',
            
            # Code
            'py': 'code',
            'js': 'code',
            'ts': 'code',
            'jsx': 'code',
            'tsx': 'code',
            'java': 'code',
            'cpp': 'code',
            'c': 'code',
            'h': 'code',
            'cs': 'code',
            'sql': 'code',
            'json': 'code',
            'xml': 'code',
            'yaml': 'code',
            'yml': 'code',
            'sh': 'code',
            'bash': 'code',
            
            # Images
            'jpg': 'image',
            'jpeg': 'image',
            'png': 'image',
            'gif': 'image',
            'bmp': 'image',
            'webp': 'image',
            'svg': 'image',
            
            # Video
            'mp4': 'video',
            'avi': 'video',
            'mov': 'video',
            'mkv': 'video',
            'webm': 'video',
        }
        
        # Try extension first
        if ext in type_map:
            return type_map[ext]
        
        # Strategy 2: MIME type hints
        mime_lower = mime_type.lower() if mime_type else ''
        
        if 'spreadsheet' in mime_lower or 'excel' in mime_lower or mime_lower.startswith('application/vnd.ms-excel') or 'sheet' in mime_lower:
            return 'table'
        elif 'image' in mime_lower or mime_lower.startswith('image/'):
            return 'image'
        elif 'video' in mime_lower or mime_lower.startswith('video/'):
            return 'video'
        elif 'json' in mime_lower or mime_lower == 'application/json':
            return 'code'
        elif 'text' in mime_lower:
            return 'document'
        
        # Default to document for unknown types (safest option)
        return 'document'
    
    @staticmethod
    def _detect_file_type_by_extension(filename: str) -> str:
        """Legacy method - use _detect_file_type_with_fallback instead."""
        # Redirect to new method
        return AttachmentService._detect_file_type_with_fallback(filename, '')

    @staticmethod
    async def get_message_attachment_contents(
        db: AsyncSession, message_id: UUID
    ) -> list[str | dict]:
        """Get all attachment contents for a message."""
        stmt = select(Attachment).where(Attachment.message_id == message_id)
        result = await db.execute(stmt)
        attachments = result.scalars().all()
        
        contents = []
        for attachment in attachments:
            content = AttachmentService.get_attachment_content(attachment)
            contents.append(content)
        
        return contents
